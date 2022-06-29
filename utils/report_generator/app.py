from openpyxl import load_workbook, utils


class ReportGenerator:

  def __init__(self, report_path: str, columns_analytical: list, columns_synthetic: list, bot_name: str):
    self.template_path = r'utils\report_generator\files\report.xlsx'
    self.bot_name = bot_name
    self.report_path = report_path
    self.columns_analytical = columns_analytical
    self.columns_synthetic = columns_synthetic
    self.__columns_interval = utils.cell.get_column_interval(1, 50)


  def initialize_report(self):
    report_path = self.report_path

    self.wb = load_workbook(self.template_path)
    self.analytical = self.wb['analitico']
    self.synthetic = self.wb['sintetico']
    self.analytical['C4'].value = self.bot_name
    self.synthetic['C4'].value = self.bot_name

    for index, column in enumerate(self.columns_analytical):
      coordenate = '{}5'.format(self.__columns_interval[index])
      self.analytical[coordenate].value = column

    for index, column in enumerate(self.columns_synthetic):
      coordenate = '{}5'.format(self.__columns_interval[index])
      self.synthetic[coordenate].value = column

    self.wb.save(report_path)
  
  def write_analytical(self, columns: object, save_report = True):
    row = self.__find_next_row(self.analytical) + 1

    for item in columns:
      column = self.__find_column(item, self.analytical)
      coordenate = '{}{}'.format(column, row)
      
      self.analytical[coordenate].value = columns[item]

      if save_report:
        self.wb.save(self.report_path)

  def write_synthetic(self, columns: object, save_report = True):
    for item in columns:
      column = self.__find_column(item, self.synthetic)
      row = self.__find_next_row(self.synthetic)
      coordenate = '{}{}'.format(column, row)
      
      self.synthetic[coordenate].value = columns[item]

      if save_report:
        self.wb.save(self.report_path)

  def write_analytical_list(self, columns_list: list):
    for item in columns_list:
      self.write_analytical(item, save_report=False)

    self.wb.save(self.report_path)

  def write_synthetic_list(self, columns_list: list):
    for item in columns_list:
      self.write_synthetic(item, save_report=False)

    self.wb.save(self.report_path)

  def save(self):
    self.wb.save(self.report_path)

  # Private methods

  def __find_column(self, column_name: str, ws):
    for letter in self.__columns_interval:
      coordenate = '{}5'.format(letter)
      if ws[coordenate].value == column_name:
        return letter
    
    raise Exception('Coluna \'{}\' não existe no relatório {}'.format(column_name, ws.title))
  
  def __find_next_row(self, ws):
    length_array = []
    for letter in self.__columns_interval:
      length_array.append(len(ws[letter]))

    return max(length_array)
